from openpyxl import load_workbook

# --근무시간표 정렬-- 2024-12-31 추가(김주상)
def sortStudentTimeSchecule(excel_file_path):
    load_wb = load_workbook(excel_file_path, data_only=True)
    work_space = load_wb.active

    # 해당 문자열이면 cell이 비였다고 판단
    empty = [None, "", " "]

    for row_start in range(11, 83, 8):
        for col_start in range(2, 22, 2):
            data = []
            # 16칸에 대한 모든 근무자 정보를 data에 넣고 기존 값 삭제
            for row in range(row_start, row_start+8):
                for col in range(col_start, col_start+2):
                    cell_value = work_space.cell(row=row, column=col).value
                    if cell_value not in empty:
                        data.append(cell_value)
                    work_space.cell(row=row, column=col).value = None
            # data에 있는 값을 위에서부터 채움
            data_index = 0
            for row in range(row_start, row_start+8):
                for col in range(col_start, col_start+2):
                    if data_index < len(data):
                        work_space.cell(row=row, column=col).value = data[data_index]
                        data_index += 1

    load_wb.save(excel_file_path)