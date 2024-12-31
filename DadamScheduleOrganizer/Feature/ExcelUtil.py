from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from Model.CellPattern import cellPattern

def findLastRowIndex(work_space, col, row):
    last_row = row
    while work_space[f"{chr(last_row)}{col}"].value is not None:
        last_row += 1

    return last_row

def findLastColIndex(work_space, col, row):
    last_col = col
    while work_space[f"{chr(row)}{last_col}"].value is not None:
        last_col += 1

    return last_col

def setMajorColor(workbook, studentName, studentMajor):
    fill = None

    if "전기" in studentMajor:
        fill = cellPattern["전전통"]

    elif "기" in studentMajor:
        fill = cellPattern["기계"]

    elif "메" in studentMajor:
        fill = cellPattern["메카"]

    elif "전" in studentMajor or "통" in studentMajor:
        fill = cellPattern["전전통"]

    elif "컴" in studentMajor:
        fill = cellPattern["컴공"]

    elif "디" in studentMajor or "건" in studentMajor:
        fill = cellPattern["디건"]

    elif "에" in studentMajor:
        fill = cellPattern["에신화"]

    elif "산" in studentMajor:
        fill = cellPattern["산경"]

    elif "용" in studentMajor:
        fill = cellPattern["고용"]

    for index in range(len(workbook.sheetnames)):
        work_space = workbook.worksheets[index]
        max_row = work_space.max_row
        max_col = work_space.max_column

        max_col_letter = get_column_letter(max_col)
        rule = CellIsRule(operator="equal", formula=[f"\"{studentName}\""], fill=fill)
        work_space.conditional_formatting.add(f"A1:{max_col_letter}{max_row}", rule)

