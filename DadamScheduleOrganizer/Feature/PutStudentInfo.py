from openpyxl import load_workbook

from Feature.ExcelUtil import findLastRowIndex, findLastColIndex
from Model.Student import Worker

def putStudentNameMajor(excel_file_path, student):
    load_wb = load_workbook(excel_file_path, data_only=True)
    work_space = load_wb.active

    if student.worker == Worker.EXISTING:
        last_row = findLastRowIndex(work_space, 4, 67)
        work_space[f"{chr(last_row)}4"] = student.name

    elif student.worker == Worker.NEW:
        last_row = findLastRowIndex(work_space, 6, 67)
        work_space[f"{chr(last_row)}6"] = student.name

    load_wb.save(excel_file_path)


def putStudnetTimeSchedule(excel_file_path, student):
    load_wb = load_workbook(excel_file_path, data_only=True)
    work_space = load_wb.active

    if student.worker == Worker.NEW:
        first_row = 68
        second_row = 69
        # 아스키 코드 D, E

    elif student.worker == Worker.EXISTING:
        first_row = 66
        second_row = 67
        # 아스키 코드 B, C

    for week_time_table in student.timeTable:
        # 월 화 수 목 금
        time_col = 11  # B11, C11, F11, G11 부터 시작
        for day_index, day_time_table in enumerate(week_time_table):
            if day_time_table != "":

                last_first_col = findLastColIndex(work_space, time_col, first_row)
                last_second_col = findLastColIndex(work_space, time_col, second_row)

                if last_first_col == last_second_col:
                    work_space[f"{chr(first_row)}{last_first_col}"] = student.name
                else:
                    work_space[f"{chr(second_row)}{last_second_col}"] = student.name
            time_col += 8

        first_row += 4
        second_row += 4
    load_wb.save(excel_file_path)

